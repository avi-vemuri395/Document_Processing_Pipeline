"""
Export functionality for extracted data.
Supports Excel and CSV formats for easy data analysis.
Part of Phase 3 implementation.
"""

import csv
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter


class DataExporter:
    """
    Export extracted data to various formats.
    Supports Excel, CSV, and JSON formats.
    """
    
    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize the exporter.
        
        Args:
            output_dir: Directory for export files
        """
        self.output_dir = output_dir or Path("outputs/exports")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(self,
                       data: Union[Dict, List[Dict]],
                       filename: str,
                       include_metadata: bool = True,
                       include_tables: bool = True) -> Path:
        """
        Export data to Excel format with formatting.
        
        Args:
            data: Extracted data (single result or list of results)
            filename: Output filename (without extension)
            include_metadata: Include metadata sheet
            include_tables: Include extracted tables
            
        Returns:
            Path to exported file
        """
        output_path = self.output_dir / f"{filename}.xlsx"
        
        # Convert single result to list
        if isinstance(data, dict):
            data = [data]
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create main data sheet
        self._create_main_sheet(wb, data)
        
        # Create summary sheet
        self._create_summary_sheet(wb, data)
        
        # Create metadata sheet if requested
        if include_metadata:
            self._create_metadata_sheet(wb, data)
        
        # Create tables sheet if requested and tables exist
        if include_tables:
            self._create_tables_sheet(wb, data)
        
        # Create comparison sheet if multiple documents
        if len(data) > 1:
            self._create_comparison_sheet(wb, data)
        
        # Save workbook
        wb.save(output_path)
        
        print(f"✅ Exported to Excel: {output_path}")
        return output_path
    
    def export_to_csv(self,
                     data: Union[Dict, List[Dict]],
                     filename: str,
                     flatten: bool = True) -> Path:
        """
        Export data to CSV format.
        
        Args:
            data: Extracted data
            filename: Output filename (without extension)
            flatten: Flatten nested structures
            
        Returns:
            Path to exported file
        """
        output_path = self.output_dir / f"{filename}.csv"
        
        # Convert single result to list
        if isinstance(data, dict):
            data = [data]
        
        # Prepare rows for CSV
        rows = []
        for result in data:
            if 'extracted_fields' in result:
                row = self._flatten_fields(result['extracted_fields']) if flatten else result['extracted_fields']
            else:
                row = self._flatten_dict(result) if flatten else result
            
            # Add metadata
            row['_document'] = result.get('document', 'unknown')
            row['_form_id'] = result.get('form_id', 'unknown')
            row['_timestamp'] = result.get('timestamp', datetime.now().isoformat())
            
            rows.append(row)
        
        # Write CSV
        if rows:
            fieldnames = set()
            for row in rows:
                fieldnames.update(row.keys())
            fieldnames = sorted(list(fieldnames))
            
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
        
        print(f"✅ Exported to CSV: {output_path}")
        return output_path
    
    def export_to_json(self,
                      data: Union[Dict, List[Dict]],
                      filename: str,
                      pretty: bool = True) -> Path:
        """
        Export data to JSON format.
        
        Args:
            data: Extracted data
            filename: Output filename (without extension)
            pretty: Pretty print JSON
            
        Returns:
            Path to exported file
        """
        output_path = self.output_dir / f"{filename}.json"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(data, f, indent=2, default=str)
            else:
                json.dump(data, f, default=str)
        
        print(f"✅ Exported to JSON: {output_path}")
        return output_path
    
    def export_multi_bank_comparison(self,
                                    results: List[Dict],
                                    filename: str = "multi_bank_comparison") -> Path:
        """
        Export multi-bank extraction results for comparison.
        
        Args:
            results: List of extraction results from different banks
            filename: Output filename
            
        Returns:
            Path to Excel file
        """
        output_path = self.output_dir / f"{filename}.xlsx"
        
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            # Create comparison DataFrame
            comparison_data = []
            
            for result in results:
                bank_data = {
                    'Bank': result.get('form_id', 'Unknown'),
                    'Document': result.get('document', 'Unknown'),
                    'Fields Extracted': result.get('metrics', {}).get('extracted_fields', 0),
                    'Total Fields': result.get('metrics', {}).get('total_fields', 0),
                    'Coverage %': result.get('metrics', {}).get('coverage_percentage', 0),
                    'Processing Time': result.get('metrics', {}).get('processing_time', 0)
                }
                
                # Add key field values
                fields = result.get('extracted_fields', {})
                for key_field in ['Name', 'Social Security Number', 'Email', 'Phone', 
                                 'Total Assets', 'Total Liabilities', 'Net Worth']:
                    bank_data[key_field] = fields.get(key_field, '')
                
                comparison_data.append(bank_data)
            
            # Write comparison sheet
            df_comparison = pd.DataFrame(comparison_data)
            df_comparison.to_excel(writer, sheet_name='Bank Comparison', index=False)
            
            # Format the comparison sheet
            workbook = writer.book
            worksheet = writer.sheets['Bank Comparison']
            
            # Add formatting
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#366092',
                'font_color': 'white',
                'border': 1
            })
            
            # Write headers with format
            for col_num, value in enumerate(df_comparison.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Add individual bank sheets
            for result in results:
                bank_name = result.get('form_id', 'Unknown').replace('_', ' ').title()
                fields = result.get('extracted_fields', {})
                
                # Create DataFrame for this bank
                bank_df = pd.DataFrame([
                    {'Field': k, 'Value': v} 
                    for k, v in fields.items()
                ])
                
                if not bank_df.empty:
                    bank_df.to_excel(writer, sheet_name=bank_name[:31], index=False)  # Excel sheet name limit
        
        print(f"✅ Exported multi-bank comparison: {output_path}")
        return output_path
    
    def _create_main_sheet(self, wb: Workbook, data: List[Dict]) -> None:
        """Create main data sheet with all extracted fields."""
        ws = wb.create_sheet("Extracted Data")
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Collect all field names
        all_fields = set()
        for result in data:
            if 'extracted_fields' in result:
                all_fields.update(result['extracted_fields'].keys())
        
        field_names = sorted(list(all_fields))
        
        # Write headers
        headers = ['Document', 'Form ID', 'Timestamp'] + field_names
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_idx, result in enumerate(data, 2):
            # Document info
            ws.cell(row=row_idx, column=1, value=result.get('document', 'unknown'))
            ws.cell(row=row_idx, column=2, value=result.get('form_id', 'unknown'))
            ws.cell(row=row_idx, column=3, value=result.get('timestamp', ''))
            
            # Field values
            fields = result.get('extracted_fields', {})
            for col_idx, field_name in enumerate(field_names, 4):
                value = fields.get(field_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    def _create_summary_sheet(self, wb: Workbook, data: List[Dict]) -> None:
        """Create summary sheet with metrics."""
        ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
        
        # Style definitions
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        
        # Title
        ws.cell(row=1, column=1, value="Extraction Summary").font = title_font
        
        # Summary data
        row = 3
        ws.cell(row=row, column=1, value="Metric").font = header_font
        ws.cell(row=row, column=2, value="Value").font = header_font
        
        # Calculate summary metrics
        total_docs = len(data)
        total_fields_extracted = sum(
            len(result.get('extracted_fields', {})) for result in data
        )
        avg_coverage = sum(
            result.get('metrics', {}).get('coverage_percentage', 0) for result in data
        ) / max(total_docs, 1)
        
        # Write metrics
        metrics = [
            ("Documents Processed", total_docs),
            ("Total Fields Extracted", total_fields_extracted),
            ("Average Coverage %", f"{avg_coverage:.1f}%"),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        for idx, (metric, value) in enumerate(metrics, 4):
            ws.cell(row=idx, column=1, value=metric)
            ws.cell(row=idx, column=2, value=value)
        
        # Add per-document summary
        if data:
            row = len(metrics) + 6
            ws.cell(row=row, column=1, value="Per-Document Summary").font = title_font
            
            row += 2
            headers = ["Document", "Form", "Fields Extracted", "Coverage %", "Tables Found"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = header_font
            
            for result in data:
                row += 1
                ws.cell(row=row, column=1, value=result.get('document', 'unknown'))
                ws.cell(row=row, column=2, value=result.get('form_id', 'unknown'))
                ws.cell(row=row, column=3, value=len(result.get('extracted_fields', {})))
                ws.cell(row=row, column=4, value=f"{result.get('metrics', {}).get('coverage_percentage', 0):.1f}%")
                ws.cell(row=row, column=5, value=len(result.get('tables', [])))
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    def _create_metadata_sheet(self, wb: Workbook, data: List[Dict]) -> None:
        """Create metadata sheet with extraction details."""
        ws = wb.create_sheet("Metadata")
        
        # Write metadata for each document
        row = 1
        for idx, result in enumerate(data):
            if idx > 0:
                row += 2  # Skip line between documents
            
            # Document header
            ws.cell(row=row, column=1, value=f"Document {idx + 1}").font = Font(bold=True)
            row += 1
            
            # Metadata fields
            metadata = result.get('metadata', {})
            for key, value in metadata.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            
            # Metrics
            metrics = result.get('metrics', {})
            if metrics:
                row += 1
                ws.cell(row=row, column=1, value="Metrics").font = Font(bold=True)
                row += 1
                
                for key, value in metrics.items():
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
    
    def _create_tables_sheet(self, wb: Workbook, data: List[Dict]) -> None:
        """Create sheet with extracted tables."""
        tables_exist = any(result.get('tables', []) for result in data)
        
        if not tables_exist:
            return
        
        ws = wb.create_sheet("Extracted Tables")
        
        row = 1
        for doc_idx, result in enumerate(data):
            tables = result.get('tables', [])
            
            if not tables:
                continue
            
            # Document header
            ws.cell(row=row, column=1, value=f"Document: {result.get('document', 'unknown')}").font = Font(bold=True, size=12)
            row += 2
            
            for table_idx, table in enumerate(tables):
                # Table header
                table_type = table.get('type', 'unknown')
                ws.cell(row=row, column=1, value=f"Table {table_idx + 1}: {table_type}").font = Font(bold=True)
                ws.cell(row=row, column=3, value=f"Page {table.get('page', '?')}")
                row += 1
                
                # Table data
                table_data = table.get('data', [])
                if table_data:
                    # Write headers
                    if isinstance(table_data[0], dict):
                        headers = list(table_data[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                        row += 1
                        
                        # Write data
                        for data_row in table_data:
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=row, column=col, value=str(data_row.get(header, '')))
                            row += 1
                
                # Extracted values
                if 'extracted_values' in table:
                    row += 1
                    ws.cell(row=row, column=1, value="Extracted Values:").font = Font(italic=True)
                    row += 1
                    
                    for key, value in table['extracted_values'].items():
                        ws.cell(row=row, column=1, value=key)
                        ws.cell(row=row, column=2, value=str(value))
                        row += 1
                
                row += 2  # Space between tables
    
    def _create_comparison_sheet(self, wb: Workbook, data: List[Dict]) -> None:
        """Create comparison sheet for multiple documents."""
        ws = wb.create_sheet("Comparison")
        
        # Get common fields
        all_fields = set()
        for result in data:
            if 'extracted_fields' in result:
                all_fields.update(result['extracted_fields'].keys())
        
        field_names = sorted(list(all_fields))
        
        # Write headers
        ws.cell(row=1, column=1, value="Field").font = Font(bold=True)
        for col, result in enumerate(data, 2):
            doc_name = result.get('document', f'Doc{col-1}')
            ws.cell(row=1, column=col, value=doc_name).font = Font(bold=True)
        
        # Write field comparisons
        for row, field_name in enumerate(field_names, 2):
            ws.cell(row=row, column=1, value=field_name)
            
            for col, result in enumerate(data, 2):
                fields = result.get('extracted_fields', {})
                value = fields.get(field_name, '')
                ws.cell(row=row, column=col, value=str(value))
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    def _flatten_dict(self, d: Dict, parent_key: str = '', sep: str = '_') -> Dict:
        """Flatten nested dictionary."""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                items.append((new_key, str(v)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    def _flatten_fields(self, fields: Dict) -> Dict:
        """Flatten field dictionary for CSV export."""
        flat = {}
        for key, value in fields.items():
            if isinstance(value, (dict, list)):
                flat[key] = str(value)
            else:
                flat[key] = value
        return flat
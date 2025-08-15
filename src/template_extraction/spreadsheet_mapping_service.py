"""
Spreadsheet Mapping Service - Populates Excel templates with master JSON data
Part of the Two-Part Pipeline (Part 2b - Spreadsheet Generation)
"""

import json
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import load_workbook


class SpreadsheetMappingService:
    """
    Maps master JSON data from Part 1 to Excel spreadsheet templates.
    
    This service takes the comprehensive extraction and populates
    various Excel templates (Debt Schedule, Use of Funds, etc.)
    """
    
    # Define available spreadsheet templates
    SPREADSHEET_TEMPLATES = {
        "debt_schedule": {
            "template": "templates/Debt Schedule Template.xlsx",
            "sheet_name": "Debt Schedule",
            "output_name": "debt_schedule_filled.xlsx"
        },
        "use_of_funds": {
            "template": "templates/Use of Funds.xlsx", 
            "sheet_name": 0,  # First sheet
            "output_name": "use_of_funds_filled.xlsx"
        },
        "projection": {
            "template": "templates/projection template.xlsx",
            "sheet_name": 0,
            "output_name": "projections_filled.xlsx"
        }
    }
    
    def __init__(self):
        """Initialize the spreadsheet mapping service"""
        self.output_base = Path("outputs/applications")
        
    def populate_all_spreadsheets(self, application_id: str) -> Dict[str, Any]:
        """
        Populate all available spreadsheet templates with master data.
        
        Args:
            application_id: Unique application identifier
            
        Returns:
            Dictionary with results for each spreadsheet
        """
        # Load master data from Part 1
        master_data = self._load_master_data(application_id)
        
        if not master_data:
            print("❌ No master data found. Run Part 1 first.")
            return {}
        
        print(f"\n{'='*70}")
        print(f"  SPREADSHEET POPULATION")
        print(f"  Application ID: {application_id}")
        print(f"{'='*70}")
        
        results = {}
        
        # Process each spreadsheet template
        for sheet_type, config in self.SPREADSHEET_TEMPLATES.items():
            print(f"\n📊 Processing {sheet_type}...")
            
            try:
                result = self._populate_spreadsheet(
                    sheet_type,
                    config,
                    master_data,
                    application_id
                )
                results[sheet_type] = result
                print(f"  ✅ Generated {config['output_name']}")
            except Exception as e:
                print(f"  ❌ Error: {e}")
                results[sheet_type] = {"status": "error", "error": str(e)}
        
        # Save summary
        self._save_summary(application_id, results)
        
        return results
    
    def populate_debt_schedule(
        self, 
        master_data: Dict[str, Any],
        application_id: str
    ) -> Path:
        """
        Populate the Debt Schedule template specifically.
        
        Args:
            master_data: Extracted data from Part 1
            application_id: Application identifier
            
        Returns:
            Path to filled spreadsheet
        """
        template_path = Path(self.SPREADSHEET_TEMPLATES["debt_schedule"]["template"])
        
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        # Load workbook
        wb = load_workbook(template_path)
        ws = wb["Debt Schedule"]
        
        # Fill applicant name (Cell D2)
        applicant_name = self._get_value(master_data, ["personal_info", "name"]) or \
                        self._get_value(master_data, ["business_info", "business_name"]) or \
                        "Unknown Applicant"
        ws['D2'] = applicant_name
        
        # Fill debt rows (starting at row 6)
        debts = self._extract_debts_from_master(master_data)
        start_row = 6
        
        for i, debt in enumerate(debts[:10], start=0):  # Limit to 10 debts
            row = start_row + i
            
            # Column C: Creditor
            if debt.get("creditor"):
                ws[f'C{row}'] = debt["creditor"]
            
            # Column D: Original Amount
            if debt.get("original_amount"):
                ws[f'D{row}'] = self._parse_number(debt["original_amount"])
            
            # Column E: Original Date
            if debt.get("original_date"):
                ws[f'E{row}'] = debt["original_date"]
            
            # Column F: Current Balance
            if debt.get("current_balance"):
                ws[f'F{row}'] = self._parse_number(debt["current_balance"])
            
            # Column G: Interest Rate
            if debt.get("interest_rate"):
                ws[f'G{row}'] = self._parse_number(debt["interest_rate"])
            
            # Column H: Monthly Payment
            if debt.get("monthly_payment"):
                ws[f'H{row}'] = self._parse_number(debt["monthly_payment"])
            
            # Column I: Maturity Date
            if debt.get("maturity_date"):
                ws[f'I{row}'] = debt["maturity_date"]
            
            # Column L: Purpose/Use
            if debt.get("purpose"):
                ws[f'L{row}'] = debt["purpose"]
        
        # Save the filled spreadsheet
        output_path = self._get_output_path(application_id, "debt_schedule_filled.xlsx")
        wb.save(output_path)
        
        return output_path
    
    def populate_use_of_funds(
        self,
        master_data: Dict[str, Any],
        application_id: str
    ) -> Path:
        """
        Populate the Use of Funds template.
        
        Args:
            master_data: Extracted data from Part 1
            application_id: Application identifier
            
        Returns:
            Path to filled spreadsheet
        """
        template_path = Path(self.SPREADSHEET_TEMPLATES["use_of_funds"]["template"])
        
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        # Load workbook
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Map master data to Use of Funds categories
        use_of_funds = self._extract_use_of_funds(master_data)
        
        # Define row mappings for each use category
        row_mappings = {
            "purchase_building": 2,
            "new_construction": 3,
            "building_improvements": 4,
            "leasehold_improvements": 5,
            "new_equipment": 6,
            "furniture_fixtures": 7,
            "working_capital": 8,
            "inventory": 9,
            "business_acquisition": 10,
            "partner_buyout": 11,
            "refinance_real_estate": 12,
            "refinance_other_debt": 13
        }
        
        # Fill in the values
        for category, row in row_mappings.items():
            if category in use_of_funds:
                data = use_of_funds[category]
                
                # Column B: Amount Already Injected
                if data.get("already_injected"):
                    ws[f'B{row}'] = self._parse_number(data["already_injected"])
                
                # Column C: Loan Amount Needed
                if data.get("loan_needed"):
                    ws[f'C{row}'] = self._parse_number(data["loan_needed"])
                
                # Column D: Total Project Cost
                if data.get("total_cost"):
                    ws[f'D{row}'] = self._parse_number(data["total_cost"])
        
        # Save the filled spreadsheet
        output_path = self._get_output_path(application_id, "use_of_funds_filled.xlsx")
        wb.save(output_path)
        
        return output_path
    
    def _populate_spreadsheet(
        self,
        sheet_type: str,
        config: Dict[str, Any],
        master_data: Dict[str, Any],
        application_id: str
    ) -> Dict[str, Any]:
        """
        Generic method to populate a spreadsheet based on type.
        
        Args:
            sheet_type: Type of spreadsheet (debt_schedule, use_of_funds, etc.)
            config: Configuration for the spreadsheet
            master_data: Data to populate
            application_id: Application identifier
            
        Returns:
            Result dictionary with status and path
        """
        if sheet_type == "debt_schedule":
            output_path = self.populate_debt_schedule(master_data, application_id)
        elif sheet_type == "use_of_funds":
            output_path = self.populate_use_of_funds(master_data, application_id)
        else:
            # For now, just copy the template for unsupported types
            template_path = Path(config["template"])
            output_path = self._get_output_path(application_id, config["output_name"])
            
            if template_path.exists():
                wb = load_workbook(template_path)
                wb.save(output_path)
            else:
                raise FileNotFoundError(f"Template not found: {template_path}")
        
        return {
            "status": "success",
            "output_path": str(output_path),
            "template": config["template"]
        }
    
    def _extract_debts_from_master(self, master_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Extract debt information from master data.
        
        Args:
            master_data: Master data from Part 1
            
        Returns:
            List of debt dictionaries
        """
        debts = []
        
        # Look in debt_schedules category
        debt_data = master_data.get("debt_schedules", {})
        
        for key, value in debt_data.items():
            if isinstance(value, dict):
                # It's a structured debt entry
                debts.append(value)
            elif "loan" in key.lower() or "debt" in key.lower():
                # Try to parse it as a debt entry
                debt_entry = {
                    "creditor": key,
                    "current_balance": value
                }
                debts.append(debt_entry)
        
        # Also look in financial_data for debt-related fields
        financial_data = master_data.get("financial_data", {})
        
        for key, value in financial_data.items():
            if any(term in key.lower() for term in ["mortgage", "loan", "credit", "debt"]):
                if key not in [d.get("creditor", "") for d in debts]:
                    debts.append({
                        "creditor": key,
                        "current_balance": value
                    })
        
        return debts
    
    def _extract_use_of_funds(self, master_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract use of funds information from master data.
        
        Args:
            master_data: Master data from Part 1
            
        Returns:
            Dictionary of use of funds categories and amounts
        """
        use_of_funds = {}
        
        # Look for relevant data in business_info and financial_data
        business_data = master_data.get("business_info", {})
        financial_data = master_data.get("financial_data", {})
        
        # Map common fields to use of funds categories
        field_mappings = {
            "working_capital": ["working_capital", "working_capital_needed"],
            "inventory": ["inventory", "inventory_value"],
            "new_equipment": ["equipment", "equipment_purchase"],
            "refinance_real_estate": ["real_estate_debt", "mortgage_refinance"],
            "business_acquisition": ["acquisition_cost", "business_purchase"]
        }
        
        for category, field_names in field_mappings.items():
            for field in field_names:
                # Check business_info
                if field in business_data:
                    use_of_funds[category] = {
                        "loan_needed": business_data[field]
                    }
                    break
                # Check financial_data
                if field in financial_data:
                    use_of_funds[category] = {
                        "loan_needed": financial_data[field]
                    }
                    break
        
        return use_of_funds
    
    def _get_value(self, data: Dict[str, Any], path: List[str]) -> Any:
        """
        Get a value from nested dictionary using a path.
        
        Args:
            data: Dictionary to search
            path: List of keys to traverse
            
        Returns:
            Value at path or None
        """
        current = data
        for key in path:
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                return None
        return current
    
    def _parse_number(self, value: Any) -> float:
        """
        Parse a value to a number, handling various formats.
        
        Args:
            value: Value to parse
            
        Returns:
            Numeric value or 0
        """
        if isinstance(value, (int, float)):
            return value
        
        if isinstance(value, str):
            # Remove common formatting
            cleaned = value.replace("$", "").replace(",", "").replace("%", "").strip()
            try:
                return float(cleaned)
            except:
                return 0
        
        return 0
    
    def _get_output_path(self, application_id: str, filename: str) -> Path:
        """
        Get the output path for a filled spreadsheet.
        
        Args:
            application_id: Application identifier
            filename: Name of output file
            
        Returns:
            Path object for output
        """
        output_dir = self.output_base / application_id / "part2_spreadsheets"
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir / filename
    
    def _load_master_data(self, application_id: str) -> Dict[str, Any]:
        """
        Load master data from Part 1.
        
        Args:
            application_id: Application identifier
            
        Returns:
            Master data dictionary or empty dict
        """
        master_path = self.output_base / application_id / "part1_document_processing" / "master_data.json"
        
        if not master_path.exists():
            return {}
        
        with open(master_path, 'r') as f:
            return json.load(f)
    
    def _save_summary(self, application_id: str, results: Dict[str, Any]):
        """
        Save a summary of spreadsheet population results.
        
        Args:
            application_id: Application identifier
            results: Results from spreadsheet population
        """
        summary = {
            "application_id": application_id,
            "timestamp": datetime.now().isoformat(),
            "spreadsheets_processed": len(results),
            "results": results
        }
        
        summary_path = self.output_base / application_id / "part2_spreadsheets" / "summary.json"
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(summary_path, 'w') as f:
            json.dump(summary, f, indent=2)
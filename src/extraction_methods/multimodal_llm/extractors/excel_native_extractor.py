"""
Native Excel extraction that preserves formulas and structure.
Falls back to image conversion only when necessary.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Tuple
from dataclasses import dataclass
from decimal import Decimal
import re


@dataclass
class ExcelExtractionResult:
    """Result from Excel extraction."""
    document_type: str
    sheets_processed: List[str]
    data: Dict[str, Any]
    formulas_preserved: Dict[str, str]
    confidence: float
    metadata: Dict[str, Any]
    errors: List[str]
    
    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}
        if self.errors is None:
            self.errors = []


class ExcelNativeExtractor:
    """Extract data from Excel files preserving structure and formulas."""
    
    def __init__(self):
        """Initialize Excel extractor."""
        self.financial_patterns = self._build_financial_patterns()
        self.debt_patterns = self._build_debt_patterns()
    
    def _build_financial_patterns(self) -> Dict[str, List[re.Pattern]]:
        """Build patterns for financial statement detection."""
        return {
            'balance_sheet': [
                re.compile(r'balance\s+sheet', re.I),
                re.compile(r'statement\s+of\s+financial\s+position', re.I),
                re.compile(r'assets.*liabilities.*equity', re.I | re.S),
            ],
            'profit_loss': [
                re.compile(r'profit\s+(?:and|&)\s+loss', re.I),
                re.compile(r'income\s+statement', re.I),
                re.compile(r'p\s*&\s*l', re.I),
            ],
            'cash_flow': [
                re.compile(r'cash\s+flow', re.I),
                re.compile(r'statement\s+of\s+cash', re.I),
            ],
            'ar_aging': [
                re.compile(r'accounts?\s+receivable', re.I),
                re.compile(r'a/?r\s+aging', re.I),
            ],
            'ap_aging': [
                re.compile(r'accounts?\s+payable', re.I),
                re.compile(r'a/?p\s+aging', re.I),
            ],
        }
    
    def _build_debt_patterns(self) -> Dict[str, List[re.Pattern]]:
        """Build patterns for debt schedule detection."""
        return {
            'creditor': [
                re.compile(r'creditor|lender|bank', re.I),
                re.compile(r'loan\s+(?:name|description)', re.I),
            ],
            'balance': [
                re.compile(r'(?:current\s+)?balance|outstanding|principal', re.I),
                re.compile(r'amount\s+owed', re.I),
            ],
            'payment': [
                re.compile(r'(?:monthly\s+)?payment|installment', re.I),
                re.compile(r'pmt|payment\s+amount', re.I),
            ],
            'rate': [
                re.compile(r'(?:interest\s+)?rate|apr|%', re.I),
            ],
            'maturity': [
                re.compile(r'maturity|due\s+date|payoff', re.I),
            ],
        }
    
    def extract(self, file_path: Union[str, Path]) -> ExcelExtractionResult:
        """
        Extract data from Excel file using native pandas reading.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            ExcelExtractionResult with extracted data
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            return ExcelExtractionResult(
                document_type='unknown',
                sheets_processed=[],
                data={},
                formulas_preserved={},
                confidence=0.0,
                metadata={'error': 'File not found'},
                errors=[f'File not found: {file_path}']
            )
        
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            # Process each sheet
            all_data = {}
            sheet_types = {}
            formulas = {}
            
            for sheet_name in sheet_names:
                # Read sheet data
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # Detect sheet type
                sheet_type = self._detect_sheet_type(df, sheet_name)
                sheet_types[sheet_name] = sheet_type
                
                # Extract based on type
                if sheet_type == 'debt_schedule':
                    sheet_data = self._extract_debt_schedule(df)
                elif sheet_type in ['balance_sheet', 'profit_loss']:
                    sheet_data = self._extract_financial_statement(df, sheet_type)
                elif sheet_type in ['ar_aging', 'ap_aging']:
                    sheet_data = self._extract_aging_report(df, sheet_type)
                else:
                    sheet_data = self._extract_generic_data(df)
                
                all_data[sheet_name] = sheet_data
                
                # Try to preserve formulas (if openpyxl available)
                try:
                    formulas[sheet_name] = self._extract_formulas(file_path, sheet_name)
                except:
                    pass
            
            # Determine primary document type
            primary_type = self._determine_primary_type(sheet_types)
            
            # Calculate confidence
            confidence = self._calculate_confidence(all_data, sheet_types)
            
            return ExcelExtractionResult(
                document_type=primary_type,
                sheets_processed=sheet_names,
                data=all_data,
                formulas_preserved=formulas,
                confidence=confidence,
                metadata={
                    'sheet_types': sheet_types,
                    'file_name': file_path.name,
                    'num_sheets': len(sheet_names)
                },
                errors=[]
            )
            
        except Exception as e:
            return ExcelExtractionResult(
                document_type='unknown',
                sheets_processed=[],
                data={},
                formulas_preserved={},
                confidence=0.0,
                metadata={'error': str(e)},
                errors=[f'Extraction error: {e}']
            )
    
    def _detect_sheet_type(self, df: pd.DataFrame, sheet_name: str) -> str:
        """Detect the type of financial data in the sheet."""
        # Check sheet name first
        sheet_name_lower = sheet_name.lower()
        
        if 'debt' in sheet_name_lower or 'loan' in sheet_name_lower:
            return 'debt_schedule'
        elif 'balance' in sheet_name_lower or 'bs' in sheet_name_lower:
            return 'balance_sheet'
        elif 'p&l' in sheet_name_lower or 'pl' in sheet_name_lower or 'income' in sheet_name_lower:
            return 'profit_loss'
        elif 'cash' in sheet_name_lower:
            return 'cash_flow'
        elif 'ar' in sheet_name_lower or 'receivable' in sheet_name_lower:
            return 'ar_aging'
        elif 'ap' in sheet_name_lower or 'payable' in sheet_name_lower:
            return 'ap_aging'
        
        # Check content
        text_content = ' '.join(df.astype(str).values.flatten()[:100])
        
        # Check financial patterns
        for doc_type, patterns in self.financial_patterns.items():
            for pattern in patterns:
                if pattern.search(text_content):
                    return doc_type
        
        # Check debt patterns
        debt_indicators = 0
        for pattern_list in self.debt_patterns.values():
            for pattern in pattern_list:
                if pattern.search(text_content):
                    debt_indicators += 1
        
        if debt_indicators >= 3:
            return 'debt_schedule'
        
        return 'generic'
    
    def _extract_debt_schedule(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Extract debt schedule data."""
        result = {
            'debts': [],
            'total_debt': 0,
            'total_monthly_payment': 0
        }
        
        # Find header row
        header_row = self._find_header_row(df, self.debt_patterns)
        
        if header_row is None:
            # Try to extract without headers
            return self._extract_debt_without_headers(df)
        
        # Set headers
        df_with_headers = df.iloc[header_row:].copy()
        df_with_headers.columns = df.iloc[header_row]
        df_data = df_with_headers.iloc[1:]
        
        # Map columns
        column_mapping = self._map_debt_columns(df_with_headers.columns)
        
        # Extract each debt
        for _, row in df_data.iterrows():
            debt = {}
            
            # Creditor name
            if column_mapping.get('creditor'):
                creditor = row[column_mapping['creditor']]
                if pd.notna(creditor) and str(creditor).strip():
                    debt['creditor_name'] = str(creditor).strip()
            
            # Current balance
            if column_mapping.get('balance'):
                balance = self._parse_monetary_value(row[column_mapping['balance']])
                if balance is not None:
                    debt['current_balance'] = balance
                    result['total_debt'] += balance
            
            # Monthly payment
            if column_mapping.get('payment'):
                payment = self._parse_monetary_value(row[column_mapping['payment']])
                if payment is not None:
                    debt['monthly_payment'] = payment
                    result['total_monthly_payment'] += payment
            
            # Interest rate
            if column_mapping.get('rate'):
                rate = self._parse_percentage(row[column_mapping['rate']])
                if rate is not None:
                    debt['interest_rate'] = rate
            
            # Maturity date
            if column_mapping.get('maturity'):
                maturity = row[column_mapping['maturity']]
                if pd.notna(maturity):
                    debt['maturity_date'] = str(maturity)
            
            if debt and debt.get('creditor_name'):
                result['debts'].append(debt)
        
        return result
    
    def _extract_financial_statement(self, df: pd.DataFrame, statement_type: str) -> Dict[str, Any]:
        """Extract balance sheet or P&L data."""
        result = {
            'statement_type': statement_type,
            'line_items': {},
            'totals': {}
        }
        
        # Look for key totals
        if statement_type == 'balance_sheet':
            total_patterns = {
                'total_assets': [r'total\s+assets', r'assets\s+total'],
                'total_liabilities': [r'total\s+liabilities', r'liabilities\s+total'],
                'total_equity': [r'total\s+equity', r'shareholders?\s+equity'],
                'current_assets': [r'current\s+assets', r'total\s+current\s+assets'],
                'current_liabilities': [r'current\s+liabilities'],
            }
        else:  # profit_loss
            total_patterns = {
                'revenue': [r'(?:total\s+)?(?:revenue|sales)', r'gross\s+receipts'],
                'cogs': [r'cost\s+of\s+goods\s+sold', r'cogs', r'cost\s+of\s+sales'],
                'gross_profit': [r'gross\s+profit', r'gross\s+margin'],
                'operating_expenses': [r'(?:total\s+)?operating\s+expenses', r'opex'],
                'net_income': [r'net\s+income', r'net\s+profit', r'net\s+earnings'],
                'ebitda': [r'ebitda', r'earnings\s+before'],
            }
        
        # Search for patterns in dataframe
        for total_name, patterns in total_patterns.items():
            value = self._find_total_value(df, patterns)
            if value is not None:
                result['totals'][total_name] = value
        
        # Extract line items
        for idx, row in df.iterrows():
            # Skip empty rows
            if row.isna().all():
                continue
            
            # Look for label-value pairs
            for i in range(len(row) - 1):
                if pd.notna(row.iloc[i]) and pd.notna(row.iloc[i + 1]):
                    label = str(row.iloc[i]).strip()
                    value = self._parse_monetary_value(row.iloc[i + 1])
                    
                    if value is not None and len(label) > 2 and not label.isdigit():
                        # Clean up label
                        label = re.sub(r'[:\s]+$', '', label)
                        result['line_items'][label] = value
        
        return result
    
    def _extract_aging_report(self, df: pd.DataFrame, report_type: str) -> Dict[str, Any]:
        """Extract AR/AP aging report data."""
        result = {
            'report_type': report_type,
            'aging_buckets': {},
            'customers' if 'ar' in report_type else 'vendors': []
        }
        
        # Look for aging buckets
        bucket_patterns = {
            'current': [r'current', r'0-?30'],
            '30_days': [r'31-?60', r'30\s+days'],
            '60_days': [r'61-?90', r'60\s+days'],
            '90_days': [r'91-?120', r'90\s+days', r'over\s+90'],
            '120_plus': [r'(?:over\s+)?120\+?', r'121\+'],
        }
        
        # Find aging columns
        header_row = None
        for idx, row in df.iterrows():
            row_str = ' '.join(row.astype(str).values)
            if any(re.search(p, row_str, re.I) for patterns in bucket_patterns.values() for p in patterns):
                header_row = idx
                break
        
        if header_row is not None:
            # Extract aging data
            df_aging = df.iloc[header_row:].copy()
            df_aging.columns = df.iloc[header_row]
            
            # Map columns to buckets
            column_mapping = {}
            for bucket, patterns in bucket_patterns.items():
                for col in df_aging.columns:
                    if any(re.search(p, str(col), re.I) for p in patterns):
                        column_mapping[bucket] = col
                        break
            
            # Sum up buckets
            for bucket, col in column_mapping.items():
                total = 0
                for val in df_aging[col].iloc[1:]:
                    parsed = self._parse_monetary_value(val)
                    if parsed:
                        total += parsed
                result['aging_buckets'][bucket] = total
        
        return result
    
    def _extract_generic_data(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Extract generic tabular data."""
        result = {
            'data_type': 'generic',
            'rows': len(df),
            'columns': len(df.columns),
            'data': []
        }
        
        # Convert to list of dicts
        if len(df) > 0:
            # Try to use first row as headers
            if df.iloc[0].notna().sum() > len(df.columns) / 2:
                df.columns = df.iloc[0]
                df_data = df.iloc[1:]
            else:
                df_data = df
            
            # Convert to records
            for _, row in df_data.iterrows():
                row_dict = {}
                for col in df.columns:
                    if pd.notna(row[col]):
                        row_dict[str(col)] = str(row[col])
                if row_dict:
                    result['data'].append(row_dict)
        
        return result
    
    def _extract_debt_without_headers(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Try to extract debt data without clear headers."""
        result = {
            'debts': [],
            'total_debt': 0,
            'extraction_method': 'pattern_matching'
        }
        
        for _, row in df.iterrows():
            # Skip empty rows
            if row.isna().all():
                continue
            
            row_values = row.dropna().values
            if len(row_values) < 2:
                continue
            
            # Look for creditor name (usually first non-numeric)
            creditor = None
            balance = None
            payment = None
            
            for val in row_values:
                val_str = str(val)
                
                # Check if it's a creditor name
                if not creditor and not self._is_numeric(val_str) and len(val_str) > 2:
                    creditor = val_str
                
                # Check for balance (larger number)
                elif not balance:
                    parsed = self._parse_monetary_value(val)
                    if parsed and parsed > 1000:  # Assume balances are > $1000
                        balance = parsed
                
                # Check for payment (smaller number)
                elif not payment:
                    parsed = self._parse_monetary_value(val)
                    if parsed and parsed < balance * 0.1:  # Payment typically < 10% of balance
                        payment = parsed
            
            if creditor and balance:
                debt = {
                    'creditor_name': creditor,
                    'current_balance': balance
                }
                if payment:
                    debt['monthly_payment'] = payment
                
                result['debts'].append(debt)
                result['total_debt'] += balance
        
        return result
    
    def _find_header_row(self, df: pd.DataFrame, patterns: Dict) -> Optional[int]:
        """Find the row containing headers."""
        for idx, row in df.iterrows():
            if idx > 10:  # Headers typically in first 10 rows
                break
            
            row_str = ' '.join(row.astype(str).values)
            matches = 0
            
            for pattern_list in patterns.values():
                for pattern in pattern_list:
                    if pattern.search(row_str):
                        matches += 1
                        break
            
            if matches >= 2:  # At least 2 pattern matches
                return idx
        
        return None
    
    def _map_debt_columns(self, columns) -> Dict[str, str]:
        """Map dataframe columns to debt fields."""
        mapping = {}
        
        for col in columns:
            col_str = str(col).lower()
            
            # Creditor
            if not mapping.get('creditor'):
                if any(term in col_str for term in ['creditor', 'lender', 'bank', 'name']):
                    mapping['creditor'] = col
            
            # Balance
            if not mapping.get('balance'):
                if any(term in col_str for term in ['balance', 'outstanding', 'principal', 'amount']):
                    mapping['balance'] = col
            
            # Payment
            if not mapping.get('payment'):
                if any(term in col_str for term in ['payment', 'pmt', 'monthly', 'installment']):
                    mapping['payment'] = col
            
            # Rate
            if not mapping.get('rate'):
                if any(term in col_str for term in ['rate', 'interest', 'apr', '%']):
                    mapping['rate'] = col
            
            # Maturity
            if not mapping.get('maturity'):
                if any(term in col_str for term in ['maturity', 'due', 'payoff', 'term']):
                    mapping['maturity'] = col
        
        return mapping
    
    def _find_total_value(self, df: pd.DataFrame, patterns: List[str]) -> Optional[float]:
        """Find a total value in the dataframe using patterns."""
        for pattern_str in patterns:
            pattern = re.compile(pattern_str, re.I)
            
            for _, row in df.iterrows():
                for i, val in enumerate(row):
                    if pd.notna(val) and pattern.search(str(val)):
                        # Look for value in same row
                        for j in range(i + 1, len(row)):
                            if pd.notna(row.iloc[j]):
                                parsed = self._parse_monetary_value(row.iloc[j])
                                if parsed:
                                    return parsed
                        
                        # Look in next column same row
                        if i < len(row) - 1:
                            parsed = self._parse_monetary_value(row.iloc[i + 1])
                            if parsed:
                                return parsed
        
        return None
    
    def _parse_monetary_value(self, value: Any) -> Optional[float]:
        """Parse monetary value from various formats."""
        if pd.isna(value):
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        value_str = str(value)
        
        # Remove currency symbols and spaces
        value_str = re.sub(r'[$,\s]', '', value_str)
        
        # Handle parentheses for negatives
        if '(' in value_str and ')' in value_str:
            value_str = '-' + value_str.replace('(', '').replace(')', '')
        
        # Handle percentage (remove %)
        value_str = value_str.replace('%', '')
        
        try:
            return float(value_str)
        except:
            return None
    
    def _parse_percentage(self, value: Any) -> Optional[float]:
        """Parse percentage value."""
        parsed = self._parse_monetary_value(value)
        
        if parsed is not None:
            # If value is greater than 1, assume it's already in percentage form
            if parsed > 1:
                return parsed / 100
            return parsed
        
        return None
    
    def _is_numeric(self, value: str) -> bool:
        """Check if string represents a numeric value."""
        try:
            float(value.replace(',', '').replace('$', '').replace('%', ''))
            return True
        except:
            return False
    
    def _extract_formulas(self, file_path: Path, sheet_name: str) -> Dict[str, str]:
        """Extract formulas from Excel file using openpyxl."""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
            
            formulas = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas[cell.coordinate] = cell.value
            
            return formulas
            
        except:
            return {}
    
    def _determine_primary_type(self, sheet_types: Dict[str, str]) -> str:
        """Determine primary document type from sheet types."""
        type_counts = {}
        for sheet_type in sheet_types.values():
            if sheet_type != 'generic':
                type_counts[sheet_type] = type_counts.get(sheet_type, 0) + 1
        
        if type_counts:
            return max(type_counts, key=type_counts.get)
        
        return 'financial_workbook'
    
    def _calculate_confidence(self, data: Dict, sheet_types: Dict) -> float:
        """Calculate overall extraction confidence."""
        confidence_scores = []
        
        for sheet_name, sheet_data in data.items():
            sheet_type = sheet_types.get(sheet_name, 'generic')
            
            if sheet_type == 'debt_schedule':
                # Check if we found debts
                if sheet_data.get('debts'):
                    confidence_scores.append(0.9)
                else:
                    confidence_scores.append(0.3)
            
            elif sheet_type in ['balance_sheet', 'profit_loss']:
                # Check if we found key totals
                if sheet_data.get('totals'):
                    num_totals = len(sheet_data['totals'])
                    confidence = min(0.95, 0.5 + num_totals * 0.1)
                    confidence_scores.append(confidence)
                else:
                    confidence_scores.append(0.4)
            
            elif sheet_type != 'generic':
                confidence_scores.append(0.7)
            else:
                confidence_scores.append(0.5)
        
        if confidence_scores:
            return sum(confidence_scores) / len(confidence_scores)
        
        return 0.0